import time
from tkinter.filedialog import askdirectory
import serial
import openpyxl
from serial.tools import list_ports
from time import sleep
from pprint import pprint
import json

a = [{'221': ' 0: 0: 0.795  0: 0'},
                  {'221': ' 0: 0: 1.309  0: 1'},
                  {'221': ' 0: 0: 3.596  0: 2'},
                  {'221': ' 0: 0: 4. 13  0: 1'},
                  {'221': ' 0: 0: 5.301  0: 1'},
                  {'221': ' 0: 0: 7.664  0: 2'},
                  {'221': ' 0: 0: 9.  7  0: 2'},
                  {'221': ' 0: 0:10.370  0: 1'},
                  {'221': ' 0: 0:12.784  0: 2'}]
print(list(a[0].keys())[0])

def loadRunners():
    try:
        f = open('runners.json', 'r')
        runners = json.load(f)
        f.close()
    except FileNotFoundError:
        runners = {}
    finally:
        return runners

def loadRoutes():
    try:
        f = open('routes.json', 'r')
        routes = json.load(f)
        f.close()
    except FileNotFoundError:
        routes = {}
    finally:
        return routes

x = 1
y = 1

a = bytes('1', encoding='utf-8')
r = bytes('2', encoding='utf-8')
s = bytes('3', encoding='utf-8')
w = bytes('4', encoding='utf-8')

stats = []
runners = loadRunners()
routes = loadRoutes()

readed = True
founded = False

wb = openpyxl.Workbook()
wb.create_sheet(title = 'Первый лист', index = 0)
sheet = wb['Первый лист']

while not founded:
    for n in serial.tools.list_ports.comports():
        try:
            sr = serial.Serial(port=n.__str__()[:4], baudrate=9600)
            sr.write(a)
            data = sr.read().decode("utf-8")
            print(data)
            if data == '2':
                print('Подключен к ' + n.__str__())
                founded = True
                break
        except Exception as err:
            print(f"Unexpected {err=}, {type(err)=}")
            raise




def readSer(stats):
    print('Приложите карту для чтения')
    sr.write(r)
    sleep(1)
    while True:
        data = sr.readline().decode("utf-8")
        print(data, end='')
        if data[0] == 'n':
            stats[len(stats) - 1] = ""

        if data[0] == 'e':
            return
        # print(data, end='')

        if data[0] == "d" or data[0] == "n":
            stats[len(stats)-1] += data[1:]


# def readPack():
#     ex = True
#     global x
#     global y
#     print('Приложите карту для чтения')
#     while ex:
#         sr.write(r)
#         sleep(1)
#         response = sr.read().decode("utf-8")
#         print(response)
#         if response != '3':
#             return
#
#         while True:
#             data = sr.readline().decode("utf-8").strip()
#             print("123", data)
#             if data[0] == "e":
#                 return
#             print(data)
#
#
#         data = sr.readline().decode("utf-8").strip()
#         print(data)
#         if data == 's':
#
#            while True:
#                 data = sr.read().decode("utf-8")
#                 print(data, end = "")
#                 if data == 'e':
#                     x = 1
#                     y = y + 1
#                     print('Данные записаны!')
#                     if input('для продолжения введите любой символ ') == 'x':
#                         ex = False
#                     break
#                 else:
#                     cell = sheet.cell(row=x, column=y)
#                     cell.value = sr.readline().decode("utf-8")
#                     x = x + 1

def registr(runners, routes):
    chipNo = input('Введите номер чипа: ')
    route = input('Введите номер маршрута: ')
    if not (route in routes):
        print("Введён неверный номер маршрута")
        return
    runners.update({f'{int(chipNo)}':{'id': f'{int(chipNo)}',
                    'routeType': route,
                    'route': []},
                    'result': 'unverif'})

    sr.write(s)
    sr.write(chipNo.encode())
    while True:
        data = sr.readline().decode("utf-8")
        # inp = sr.readline().decode('utf-8')
        print(data, end='')
        if data[0] == 'd':
            print("Номер чипа:", data[1:])
        if data[0] == 'e':
            return runners

def save(runners, stats):
    with open('runners.json', 'w', encoding="windows-1251") as f:
        json.dump(runners, f)
    f.close()

    with open("stats.txt", "w", encoding="windows-1251") as txt_file:
        for line in stats:
            txt_file.write(line.replace("\r\n", "\n") + "\n")
    txt_file.close()



def addRunner(runners, data):
    data = data.split("\r\n")
    id = str(int(data[0].rsplit(" ", 1)[1]))
    if not (id in runners):
        print("Пользователь не найден.\nДобавление нового пользователя")
        route = input("Введите тип маршрута: ")
        if not (route in routes):
            print("Маршрута не существует")
            return id
        runners.update({f'{id}': {'id': f'{id}',
                        'routeType': route,
                        'route': []},
                        'result': 'unverif'})
    runners[id]['route'] = []
    for i in range(1, len(data)-1):
        kp = data[i].split(" ", 1)
        runners[id]['route'].append({f'{int(kp[0])}': kp[1]})
    return id

def verifyRoute(runner, routes):
    way = runner['route']
    if len(way) != len(routes[runner['routeType']]):
        runner['result'] = 'Failed'
        print(1)
        return 0

    for i in range(len(way)):
        if int(list(way[i].keys())[0]) != int(routes[runner['routeType']][i]):
            runner['result'] = 'Failed'
            print(list(way[i].keys())[0], routes[runner['routeType']][i])
            return 0
    runner['result'] = 'OK'
    return 1

print("Маршруты:")
pprint(routes)

while True:

    action = input('Режим работы r - регистрация, f - запись с карты, s- сохранить, c - проверить участника, x - выход в меню, v - вывод участников, vv - вывод всего введённого\n')

    if action == 'f':
        stats.append("")
        readSer(stats)
        id = addRunner(runners, stats[len(stats)-1])
        # try:
        pprint(runners[id])
        save(runners, stats)
        if(verifyRoute(runners[id], routes)):
            print('Отметка: ОК')
        else:
            print('Отметка: ПЛОХАЯ')
        # except Exception as e:
        #     print(e)
        #     pass
    elif action == 'r':
       registr(runners, routes)
       save(runners, stats)
    elif action == 'c':
        id = input("Введите номер участника: ")
        pprint(runners[id])
        if (verifyRoute(runners[id], routes)):
            print('Отметка: ОК')
        else:
            print('Отметка: ПЛОХАЯ')
    elif action == 's':
        save(runners, stats)
    elif action == 'v':
        pprint(runners)
    elif action == 'vv':
        pprint(stats)

    sr.write(b'0')




