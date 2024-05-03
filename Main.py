import time
from tkinter.filedialog import askdirectory
import serial
import openpyxl
from serial.tools import list_ports

x = 1
y = 1

a = bytes('1', encoding='utf-8')
f = bytes('f', encoding='utf-8')
r = bytes('r', encoding='utf-8')

readed = True
founded = False

wb = openpyxl.Workbook()
wb.create_sheet(title = 'Первый лист', index = 0)
sheet = wb['Первый лист']

while not founded:
    for n in list_ports.comports():
        try:
            sr = serial.Serial(port=n.__str__()[:4], baudrate=9600)
            sr.write(a)
            if sr.read().decode("Ascii") == '2':
                print('Подключен к ' + n.__str__())
                founded = True
        except Exception as err:
            print(f"Unexpected {err=}, {type(err)=}")
            raise

def readPack():
    ex = True
    global x
    global y
    print('Приложите карту для чтения')
    while ex:
                sr.write(f)
                if sr.read().decode("Ascii") == 's':

                   while True:
                        if sr.read().decode("Ascii") == 'e':
                            x = 1
                            y = y + 1
                            print('Данные записаны!')
                            if input('для продолжения введите любой символ ') == 'x':
                                ex = False
                            break
                        else:
                            cell = sheet.cell(row=x, column=y)
                            cell.value = sr.readline().decode("Ascii")
                            print(cell.value)
                            x = x + 1

def registr():
    while True:
        c = input('...')
        if c == 'x':
            break
        sr.write(r)
        sr.write(c.encode())
        print(sr.readline().decode('utf-8'))

def save():
    global dst
    dst = askdirectory()
    try:
        wb.save(dst + '/test.xlsx')
        print('Сохранено в ' + dst + '/test.xlsx')
    except:
        print('Не удалось сохранить')

while True:

    a = input('Режим работы r - регистрация, f - запись с карты, t- сохранить таблицу, x - выход в меню  ')

    if a == 'f':
        readPack()

    if a == 'r':
       registr()

    if a == 't':
        save()




