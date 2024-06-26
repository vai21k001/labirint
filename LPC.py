import time
from tkinter.filedialog import askdirectory
import serial
import openpyxl
from serial.tools import list_ports

x = 1
y = 1

a = bytes('1', encoding='utf-8')
r = bytes('2', encoding='utf-8')
s = bytes('3', encoding='utf-8')
w = bytes('4', encoding='utf-8')


readed = True
founded = False

wb = openpyxl.Workbook()
wb.create_sheet(title = 'Первый лист', index = 0)
sheet = wb['Первый лист']

while not founded:
    for n in serial.tools.list_ports.comports():
        try:
            sr = serial.Serial(port=n.__str__()[:4], baudrate=9600)
            sr.write(a)
            if sr.read().decode("Ascii") == '2':
                print('Подключен к ' + n.__str__())
                founded = True
                break
        except Exception as err:
            print(f"Unexpected {err=}, {type(err)=}")
            raise

def readPack():
    ex = True
    global x
    global y
    print('Приложите карту для чтения')
    while ex:
        sr.write(r)
        data = sr.read().decode("Ascii")
        print(data)
        if data == 's':

           while True:
                data = sr.read().decode("Ascii")
                print(data, end = "")
                if data == 'e':
                    x = 1
                    y = y + 1
                    print('Данные записаны!')
                    if input('для продолжения введите любой символ ') == 'x':
                        ex = False
                    break
                else:
                    cell = sheet.cell(row=x, column=y)
                    cell.value = sr.readline().decode("Ascii")
                    x = x + 1

def registr():
    sr.write(s)
    c = input('... ')
    sr.write(c.encode())
    while True:
        inp = sr.readline().decode('utf-8')
        if len(inp) > 2:
            print('зарегистрирован номер ' + inp)
            if input('для записи приложите карту и введите 1 ') == '1':
                sr.write(w)
            break
def save():
    global dst
    dst = askdirectory()
    try:
        wb.save(dst + '/test.xlsx')
        print('Сохранено в ' + dst + '/test.xlsx')
    except:
        print('Не удалось сохранить')

while True:

    a = input('Режим работы r - регистрация, f - запись с карты, t- сохранить таблицу, x - выход в меню  ')

    if a == 'f':
        readPack()

    if a == 'r':
       registr()

    if a == 't':
        save()




